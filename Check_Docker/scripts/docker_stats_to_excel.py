#!/usr/bin/env python3
import json, sys, pathlib
import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill

# Args:
# 1: ruta JSON consolidado
# 2: ruta Excel salida
# 3..8: umbrales (cpu_hi, cpu_md, mem_hi, mem_md, pids_hi, pids_md)
if len(sys.argv) < 3:
    print("Uso: docker_stats_to_excel.py <json_in> <xlsx_out> [cpu_hi cpu_md mem_hi mem_md pids_hi pids_md]")
    sys.exit(1)

json_path = pathlib.Path(sys.argv[1])
xlsx_path = pathlib.Path(sys.argv[2])

cpu_hi = int(sys.argv[3]) if len(sys.argv) > 3 else 85
cpu_md = int(sys.argv[4]) if len(sys.argv) > 4 else 70
mem_hi = int(sys.argv[5]) if len(sys.argv) > 5 else 90
mem_md = int(sys.argv[6]) if len(sys.argv) > 6 else 75
pids_hi = int(sys.argv[7]) if len(sys.argv) > 7 else 400
pids_md = int(sys.argv[8]) if len(sys.argv) > 8 else 200

data = json.loads(json_path.read_text(encoding="utf-8"))

def risk(row):
    r = []
    if row.get("cpu_perc",0) >= cpu_hi: r.append("CPU")
    if row.get("mem_perc",0) >= mem_hi: r.append("MEM")
    if row.get("pids",0) >= pids_hi: r.append("PIDs")
    if r: return "ALTO (" + ",".join(r) + ")"
    r = []
    if row.get("cpu_perc",0) >= cpu_md: r.append("CPU")
    if row.get("mem_perc",0) >= mem_md: r.append("MEM")
    if row.get("pids",0) >= pids_md: r.append("PIDs")
    if r: return "MEDIO (" + ",".join(r) + ")"
    return "OK"

# enriquecer
for d in data:
    d["mem_used_mb"]  = round(d.get("mem_used_bytes",0)/1024/1024, 1)
    lim = d.get("mem_limit_bytes",0)
    d["mem_limit_mb"] = round(lim/1024/1024, 1) if lim else None
    d["net_rx_mb"]    = round(d.get("net_rx_bytes",0)/1024/1024, 1)
    d["net_tx_mb"]    = round(d.get("net_tx_bytes",0)/1024/1024, 1)
    d["blk_read_mb"]  = round(d.get("blk_read_bytes",0)/1024/1024, 1)
    d["blk_write_mb"] = round(d.get("blk_write_bytes",0)/1024/1024, 1)
    d["riesgo"] = risk(d)

df = pd.DataFrame(data)
if df.empty:
    print("No hay datos para exportar.")
    sys.exit(0)

cols = [
  ("host","Host"),
  ("container_name","Contenedor"),
  ("container_id","ID"),
  ("cpu_perc","CPU %"),
  ("mem_perc","Mem %"),
  ("mem_used_mb","Mem usada (MB)"),
  ("mem_limit_mb","Límite Mem (MB)"),
  ("pids","PIDs"),
  ("net_rx_mb","Net RX (MB)"),
  ("net_tx_mb","Net TX (MB)"),
  ("blk_read_mb","Blk Read (MB)"),
  ("blk_write_mb","Blk Write (MB)"),
  ("riesgo","Riesgo"),
]
df = df[[c for c,_ in cols]]
df.columns = [n for _,n in cols]

resumen = (df.groupby(["Host","Riesgo"])
             .size()
             .unstack(fill_value=0)
             .reset_index()
             .sort_values("Host"))

with pd.ExcelWriter(xlsx_path, engine="openpyxl") as xw:
    df.to_excel(xw, sheet_name="Contenedores", index=False)
    resumen.to_excel(xw, sheet_name="Resumen", index=False)

# Formato condicional simple
wb = load_workbook(xlsx_path)
ws = wb["Contenedores"]
hdr = {cell.value: idx+1 for idx, cell in enumerate(ws[1])}
col_cpu = hdr.get("CPU %"); col_mem = hdr.get("Mem %"); col_pids = hdr.get("PIDs")

fill_red    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
fill_yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

max_row = ws.max_row
if col_cpu:
    rng = f"{ws.cell(row=2, column=col_cpu).coordinate}:{ws.cell(row=max_row, column=col_cpu).coordinate}"
    ws.conditional_formatting.add(rng, CellIsRule(operator='greaterThanOrEqual', formula=[str(cpu_hi)], fill=fill_red))
    ws.conditional_formatting.add(rng, CellIsRule(operator='between', formula=[str(cpu_md), str(cpu_hi-0.0001)], fill=fill_yellow))
if col_mem:
    rng = f"{ws.cell(row=2, column=col_mem).coordinate}:{ws.cell(row=max_row, column=col_mem).coordinate}"
    ws.conditional_formatting.add(rng, CellIsRule(operator='greaterThanOrEqual', formula=[str(mem_hi)], fill=fill_red))
    ws.conditional_formatting.add(rng, CellIsRule(operator='between', formula=[str(mem_md), str(mem_hi-0.0001)], fill=fill_yellow))
if col_pids:
    rng = f"{ws.cell(row=2, column=col_pids).coordinate}:{ws.cell(row=max_row, column=col_pids).coordinate}"
    ws.conditional_formatting.add(rng, CellIsRule(operator='greaterThanOrEqual', formula=[str(pids_hi)], fill=fill_red))
    ws.conditional_formatting.add(rng, CellIsRule(operator='between', formula=[str(pids_md), str(pids_hi-0.0001)], fill=fill_yellow))

# Auto-ancho básico
for col in ws.columns:
    max_len = 0
    letter = col[0].column_letter
    for cell in col:
        val = "" if cell.value is None else str(cell.value)
        if len(val) > max_len: max_len = len(val)
    ws.column_dimensions[letter].width = min(max_len+2, 50)

wb.save(xlsx_path)
print(f"Generado: {xlsx_path}")
