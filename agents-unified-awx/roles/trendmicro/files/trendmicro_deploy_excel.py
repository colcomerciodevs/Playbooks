#!/usr/bin/env python3
"""
Script: generate_excel.py
Objetivo:
  - Tomar el JSON consolidado generado por Ansible (con los resultados de instalación/validación)
  - Convertirlo en un archivo Excel con colores (verde si OK, naranja si REVISAR)
Uso:
  python3 generate_excel.py <input.json> <output.xlsx>
Dependencias:
  - openpyxl
"""

import json
import sys
import openpyxl  # type: ignore
from openpyxl.styles import PatternFill, Font  # type: ignore
from openpyxl.utils import get_column_letter  # type: ignore

# ===================== VALIDACIÓN DE ARGUMENTOS =====================
if len(sys.argv) < 3:
    print("Uso: python3 generate_excel.py <input.json> <output.xlsx>")
    sys.exit(1)

input_file = sys.argv[1]
output_file = sys.argv[2]

# ===================== LECTURA DEL JSON =====================
with open(input_file, encoding="utf-8") as f:
    data = json.load(f)

# Aceptar un objeto único o lista de objetos
if isinstance(data, dict):
    data = [data]

# ===================== CREACIÓN DEL EXCEL =====================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Reporte TrendMicro"

# ===================== ENCABEZADOS (actualizados) =====================
# Se agregan "Hostname" e "IP" justo después de "Host"
headers = [
    "Host",
    "Hostname",
    "IP",
    "Version Before",
    "Release Before",
    "Version After",
    "Release After",
    "tmxbc Version Before",  # NUEVO: versión del sensor antes
    "tmxbc Version After",   # NUEVO: versión del sensor después
    "tmxbc Before",          # estado inicial del sensor
    "tmxbc After",           # estado final del sensor
    "ds_agent Before",       # estado inicial del agente
    "ds_agent After",        # estado final del agente
    "Estado Final"           # OK / REVISAR
]
ws.append(headers)

# Encabezado en negrita y autofiltro
for cell in ws[1]:
    cell.font = Font(bold=True)
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
ws.freeze_panes = "A2"

# ===================== DEFINICIÓN DE COLORES =====================
green = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")   # Verde
orange = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid") # Naranja

# ===================== NORMALIZADOR =====================
def norm(v):
    """Normaliza cualquier valor a str para Excel."""
    if v is None:
        return ""
    if isinstance(v, list):
        return " ".join(str(x) for x in v)
    if isinstance(v, dict):
        return json.dumps(v, ensure_ascii=False)
    return str(v)

# ===================== ITERACIÓN SOBRE LOS HOSTS =====================
for item in data:
    services_after = item.get("services_after", {}) or {}
    services_before = item.get("services_before", {}) or {}

    # Compatibilidad y preferencia por nuevas claves del rol
    # tmxbc BEFORE (estado)
    tmxbc_before_state = item.get("tmxbc_state_before")
    if tmxbc_before_state is None:
        tmxbc_before_state = services_before.get("tmxbc", "no instalado")

    # tmxbc AFTER (estado)
    tmxbc_after_state = item.get("tmxbc_state_after")
    if tmxbc_after_state is None:
        tmxbc_after_state = services_after.get("tmxbc", "no instalado")

    # ds_agent BEFORE (estado)
    ds_before_state = item.get("ds_agent_state_before")
    if ds_before_state is None:
        ds_before_state = services_before.get("ds_agent", "no instalado")

    # ds_agent AFTER (estado)
    ds_after_state = item.get("ds_agent_state_after")
    if ds_after_state is None:
        ds_after_state = services_after.get("ds_agent", "no instalado")

    # Hostname e IP (nuevos campos)
    hostname = item.get("hostname", "")
    ip_addr = item.get("ip", "")

    # Versiones de tmxbc (nuevos campos)
    tmxbc_ver_before = item.get("tmxbc_version_before", "")
    tmxbc_ver_after = item.get("tmxbc_version_after", "")

    # Construir fila (Host -> Hostname -> IP -> versiones ds -> versiones tmxbc -> estados -> estado final)
    row = [
        norm(item.get("host", "")),
        norm(hostname),
        norm(ip_addr),
        norm(item.get("ds_agent_version_before", "")),
        norm(item.get("ds_agent_release_before", "")),
        norm(item.get("ds_agent_version_after", "")),
        norm(item.get("ds_agent_release_after", "")),
        norm(tmxbc_ver_before),
        norm(tmxbc_ver_after),
        norm(tmxbc_before_state),
        norm(tmxbc_after_state),
        norm(ds_before_state),
        norm(ds_after_state),
        norm(item.get("estado_final", "")),
    ]
    ws.append(row)

    # ===================== COLOR SOLO EN "Estado Final" =====================
    estado_cell = ws.cell(row=ws.max_row, column=len(headers))
    valor_estado = str(row[-1]).strip().upper()  # normaliza "ok", " OK ", etc.
    if valor_estado == "OK":
        estado_cell.fill = green
    else:
        estado_cell.fill = orange

# ===================== AJUSTE APROXIMADO DE ANCHOS =====================
for col_idx in range(1, len(headers) + 1):
    col_letter = get_column_letter(col_idx)
    max_len = 0
    for cell in ws[col_letter]:
        val = str(cell.value) if cell.value is not None else ""
        if len(val) > max_len:
            max_len = len(val)
    # margen mínimo y límite razonable
    ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 60)

# ===================== GUARDAR ARCHIVO =====================
wb.save(output_file)
print(f"Excel generado en {output_file}")
