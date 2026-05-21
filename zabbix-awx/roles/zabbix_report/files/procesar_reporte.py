#!/usr/bin/env python3
"""
procesar_reporte.py
-------------------
Etapa de PROCESAMIENTO del flujo Zabbix.

NO se conecta a Zabbix.

Lee el JSON producido por extraer_zabbix.py, calcula/lee promedios y máximos,
parsea umbrales, convierte unidades y genera el Excel.

Es compatible con dos formatos:

1. Formato nuevo recomendado:
    "datos": {
        "avg": 10.5,
        "max": 80.2,
        "min": 2.1,
        "ultimo": 11.0,
        "puntos": 720
    }

2. Formato viejo:
    "datos": [
        {"value_avg": "10", "value_max": "20"},
        {"value_avg": "11", "value_max": "30"}
    ]

Uso:

    python3 procesar_reporte.py \
      --input datos.json.gz \
      --output Reporte.xlsx
"""

import argparse
import gzip
import json
import re
import sys
from datetime import datetime

import pandas as pd

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def cargar_json(ruta):
    abrir = gzip.open if ruta.endswith(".gz") else open

    with abrir(ruta, "rt", encoding="utf-8") as fh:
        return json.load(fh)


def convertir_float(valor):
    try:
        return float(valor)
    except (TypeError, ValueError):
        return None


def extraer_umbrales(triggers):
    """
    Extrae umbrales desde los triggers de Zabbix.

    Retorna lista tipo:
        [
          "Memoria-Critico [> 96]",
          "Memoria-Warning [> 90]"
        ]
    """
    umbrales = []

    for trigger in triggers:
        desc = trigger.get("description", "")
        expr = trigger.get("expression", "")

        match = re.search(
            r"([><=]+)\s*(\d+(?:\.\d+)?|\{\$[^\}]+})",
            expr,
        )

        if match:
            umbrales.append(
                f"{desc} [{match.group(1)} {match.group(2)}]"
            )
        else:
            umbrales.append(desc)

    return umbrales


def reducir_serie_vieja(puntos, usar_tendencias):
    """
    Compatibilidad con el formato viejo:

        "datos": [
            {"value_avg": "...", "value_max": "..."},
            ...
        ]

    Retorna:
        promedio, maximo, minimo, ultimo, puntos
    """
    if not puntos:
        return None, None, None, None, 0

    if usar_tendencias:
        valores_avg = []
        valores_max = []
        valores_min = []

        for punto in puntos:
            avg = convertir_float(punto.get("value_avg"))
            maximo = convertir_float(punto.get("value_max"))
            minimo = convertir_float(punto.get("value_min"))

            if avg is not None:
                valores_avg.append(avg)
            if maximo is not None:
                valores_max.append(maximo)
            if minimo is not None:
                valores_min.append(minimo)

        if not valores_avg and not valores_max:
            return None, None, None, None, len(puntos)

        promedio = (
            sum(valores_avg) / len(valores_avg)
            if valores_avg else None
        )

        maximo = max(valores_max) if valores_max else None
        minimo = min(valores_min) if valores_min else None
        ultimo = valores_avg[-1] if valores_avg else None

        return promedio, maximo, minimo, ultimo, len(puntos)

    valores = []

    for punto in puntos:
        valor = convertir_float(punto.get("value"))
        if valor is not None:
            valores.append(valor)

    if not valores:
        return None, None, None, None, len(puntos)

    promedio = sum(valores) / len(valores)
    maximo = max(valores)
    minimo = min(valores)
    ultimo = valores[-1]

    return promedio, maximo, minimo, ultimo, len(puntos)


def obtener_resumen_datos(datos, usar_tendencias):
    """
    Normaliza datos nuevos o datos viejos.

    Formato nuevo:
        datos es dict con avg/max/min/ultimo/puntos.

    Formato viejo:
        datos es list y se reduce aquí.
    """
    if isinstance(datos, dict):
        promedio = convertir_float(datos.get("avg"))
        maximo = convertir_float(datos.get("max"))
        minimo = convertir_float(datos.get("min"))
        ultimo = convertir_float(datos.get("ultimo"))

        try:
            puntos = int(datos.get("puntos", 0))
        except (TypeError, ValueError):
            puntos = 0

        return promedio, maximo, minimo, ultimo, puntos

    if isinstance(datos, list):
        return reducir_serie_vieja(datos, usar_tendencias)

    return None, None, None, None, 0


def bytes_a_gb(valor):
    if valor is None:
        return None

    return valor / (1024 ** 3)


def construir_fila_metricas(maquina):
    """
    Construye una fila para la hoja principal:
        Metricas_Infraestructura

    Se quitaron:
        - Puntos CPU
        - Puntos RAM
    """
    fila = {
        "Nombre Activo": maquina.get("nombre_maquina", ""),
        "ip": maquina.get("objetivo", ""),
        "servicio": maquina.get("grupo", ""),
        "memoria actual": 0,
        "% uso memoria ram AVG": 0,
        "% uso memoria ram MAX": 0,
        "CPU actual asignada": 0,
        "% uso procesador AVG": 0,
        "% uso procesador MAX": 0,
    }

    notas_umbrales = []

    for metrica in maquina.get("metricas", []):
        concepto = metrica.get("nombre_reporte", "")
        usar_tendencias = metrica.get("usar_tendencias", False)
        datos = metrica.get("datos", {})

        promedio, maximo, minimo, ultimo, puntos = obtener_resumen_datos(
            datos,
            usar_tendencias,
        )

        if promedio is None and maximo is None and ultimo is None:
            continue

        valor_base = maximo
        if valor_base is None:
            valor_base = ultimo
        if valor_base is None:
            valor_base = promedio

        if "Number" in concepto:
            fila["CPU actual asignada"] = int(round(valor_base))

        elif "Total memory" in concepto:
            memoria_gb = bytes_a_gb(valor_base)
            if memoria_gb is not None:
                fila["memoria actual"] = f"{memoria_gb:.2f} GB"

        elif "utilization" in concepto.lower():
            if "CPU" in concepto:
                fila["% uso procesador AVG"] = (
                    f"{promedio:.4f} %" if promedio is not None else 0
                )
                fila["% uso procesador MAX"] = (
                    f"{maximo:.4f} %" if maximo is not None else 0
                )

            else:
                fila["% uso memoria ram AVG"] = (
                    f"{promedio:.4f} %" if promedio is not None else 0
                )
                fila["% uso memoria ram MAX"] = (
                    f"{maximo:.4f} %" if maximo is not None else 0
                )

            for umbral in extraer_umbrales(metrica.get("triggers", [])):
                notas_umbrales.append(f"{concepto}: {umbral}")

    fila["Umbrales detectados"] = (
        " | ".join(notas_umbrales)
        if notas_umbrales else "—"
    )

    return fila


def construir_filas_summary(maquinas):
    """
    Construye la hoja Summary por servidor y por métrica.

    Formato:
        Grupo
        Nombre maquina
        METRICA
        Promedio
        Maximo
        Umbrales detectados
    """
    filas = []

    for maquina in maquinas:
        grupo = maquina.get("grupo", "")
        nombre_maquina = maquina.get("nombre_maquina", "")
        ip = maquina.get("objetivo", "")

        metricas_utilizacion = [
            m for m in maquina.get("metricas", [])
            if "utilization" in m.get("nombre_reporte", "").lower()
        ]

        if not metricas_utilizacion:
            filas.append({
                "Grupo": grupo,
                "Nombre maquina": nombre_maquina,
                "IP": ip,
                "Metrica": "Sin métricas de utilización",
                "Promedio": "N/A",
                "Maximo": "N/A",
                "Umbrales detectados": "—",
            })
            continue

        for metrica in metricas_utilizacion:
            concepto = metrica.get("nombre_reporte", "")
            item_name = metrica.get("item_name", "")
            usar_tendencias = metrica.get("usar_tendencias", False)
            datos = metrica.get("datos", {})

            promedio, maximo, minimo, ultimo, puntos = obtener_resumen_datos(
                datos,
                usar_tendencias,
            )

            umbrales = extraer_umbrales(metrica.get("triggers", []))

            if umbrales:
                umbrales_txt = "\n".join([f"- {u}" for u in umbrales])
            else:
                umbrales_txt = "—"

            nombre_metrica = item_name if item_name else concepto

            filas.append({
                "Grupo": grupo,
                "Nombre maquina": nombre_maquina,
                "IP": ip,
                "Metrica": nombre_metrica,
                "Promedio": f"{promedio:.4f} %" if promedio is not None else "N/A",
                "Maximo": f"{maximo:.4f} %" if maximo is not None else "N/A",
                "Umbrales detectados": umbrales_txt,
            })

        filas.append({
            "Grupo": "",
            "Nombre maquina": "",
            "IP": "",
            "Metrica": "",
            "Promedio": "",
            "Maximo": "",
            "Umbrales detectados": "",
        })

    return filas


def aplicar_formato_excel(writer):
    """
    Aplica formato básico a las hojas del Excel.
    """
    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    bold_font = Font(bold=True)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for nombre_hoja in writer.sheets:
        ws = writer.sheets[nombre_hoja]

        ws.freeze_panes = "A2"

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = thin_border

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )
                cell.border = thin_border

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter

            for cell in col:
                valor = str(cell.value) if cell.value is not None else ""
                max_length = max(max_length, len(valor))

            ancho = min(max(max_length + 2, 18), 55)
            ws.column_dimensions[column].width = ancho

        if nombre_hoja == "Summary":
            ws.column_dimensions["A"].width = 18
            ws.column_dimensions["B"].width = 28
            ws.column_dimensions["C"].width = 18
            ws.column_dimensions["D"].width = 38
            ws.column_dimensions["E"].width = 18
            ws.column_dimensions["F"].width = 18
            ws.column_dimensions["G"].width = 45

            for row in ws.iter_rows(min_row=2):
                grupo = row[0].value
                nombre = row[1].value
                metrica = row[3].value

                if grupo:
                    row[0].font = bold_font
                    row[1].font = bold_font

                if metrica:
                    row[3].font = bold_font


def generar_excel(filas_metricas, filas_summary, ruta_salida, trafico):
    df_metricas = pd.DataFrame(filas_metricas)
    df_summary = pd.DataFrame(filas_summary)

    if df_metricas.empty:
        raise ValueError("No hay filas para generar el Excel.")

    resumen_general = {
        "Indicador": [
            "Máquinas procesadas",
            "Tráfico subida (MB)",
            "Tráfico bajada (MB)",
            "Generado",
        ],
        "Valor": [
            len(df_metricas),
            f"{trafico.get('subida_bytes', 0) / (1024 ** 2):.2f}",
            f"{trafico.get('bajada_bytes', 0) / (1024 ** 2):.2f}",
            datetime.now().strftime("%Y-%m-%d %H:%M"),
        ],
    }

    df_resumen_general = pd.DataFrame(resumen_general)

    with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
        df_metricas.to_excel(
            writer,
            index=False,
            sheet_name="Metricas_Infraestructura",
        )

        df_summary.to_excel(
            writer,
            index=False,
            sheet_name="Summary",
        )

        df_resumen_general.to_excel(
            writer,
            index=False,
            sheet_name="Resumen_General",
        )

        aplicar_formato_excel(writer)

    return ruta_salida


def main():
    parser = argparse.ArgumentParser(
        description="Procesa JSON de Zabbix y genera Excel"
    )

    parser.add_argument(
        "--input",
        required=True,
        help="JSON de entrada .json o .json.gz",
    )

    parser.add_argument(
        "--output",
        required=True,
        help="Ruta del Excel a generar",
    )

    args = parser.parse_args()

    payload = cargar_json(args.input)
    maquinas = payload.get("maquinas", [])
    trafico = payload.get("trafico", {})

    if not maquinas:
        print("[!] No hay máquinas en el JSON de entrada.", file=sys.stderr)
        sys.exit(1)

    filas_metricas = [
        construir_fila_metricas(maquina)
        for maquina in maquinas
    ]

    filas_summary = construir_filas_summary(maquinas)

    try:
        ruta = generar_excel(
            filas_metricas,
            filas_summary,
            args.output,
            trafico,
        )
    except Exception as exc:
        print(f"[!] Error generando Excel: {exc}", file=sys.stderr)
        sys.exit(1)

    print(f"Excel generado: {ruta} ({len(filas_metricas)} máquina(s))")


if __name__ == "__main__":
    main()