#!/usr/bin/env python3
# ============================================================
# Script: convert_json_to_excel.py
# Objetivo:
#   Leer Salidas_Playbooks/version_sistemas.json (lista de objetos)
#   y generar un Excel con columnas:
#     - inventory_name  (nombre tal cual en el inventario)
#     - hostname        (hostname real del S.O.)
#     - ip              (IP tomada del inventario / fallback)
#     - os_full_version (versión completa del S.O.)
#
# Mejoras:
#   - Nueva columna "inventory_name"
#   - Orden de columnas fijo
#   - Filtro automático y encabezado congelado
#   - Autoajuste de ancho de columnas
#
# Requisitos: pandas, openpyxl
# Uso:
#   python3 scripts/convert_json_to_excel.py
# ============================================================

import json
import os
import sys
from datetime import datetime

try:
    import pandas as pd
except Exception:
    print("ERROR: Se requiere 'pandas' instalado. Instale con: pip install pandas openpyxl", file=sys.stderr)
    sys.exit(2)

# ------------------------------------------------------------
# 1) Rutas base
#    - BASE_DIR: raíz del proyecto (carpeta padre de 'scripts')
#    - OUT_DIR : carpeta de salidas
#    - JSON_PATH: archivo de entrada (generado por el playbook)
#    - XLSX_PATH: salida Excel con timestamp
# ------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.join(BASE_DIR, "Salidas_Playbooks")
JSON_PATH = os.path.join(OUT_DIR, "version_sistemas.json")
TS = datetime.now().strftime("%Y%m%d_%H%M%S")
XLSX_PATH = os.path.join(OUT_DIR, f"version_sistemas_{TS}.xlsx")

# ------------------------------------------------------------
# 2) Función utilitaria: autoajustar ancho de columnas
#    - Calcula la longitud máxima entre encabezado y celdas
#    - Aplica un padding para respirar en la vista
# ------------------------------------------------------------
def autosize_columns(ws, df, padding=2):
    from openpyxl.utils import get_column_letter

    for idx, col in enumerate(df.columns, start=1):
        header_len = len(str(col))
        body_len = max((len(str(x)) for x in df[col].astype(str)), default=0)
        width = min(max(header_len, body_len) + padding, 80)  # cap a 80
        ws.column_dimensions[get_column_letter(idx)].width = width

def main():
    # --------------------------------------------------------
    # 3) Validar que existe el JSON de entrada
    # --------------------------------------------------------
    if not os.path.exists(JSON_PATH):
        print(f"ERROR: No existe el archivo JSON esperado: {JSON_PATH}", file=sys.stderr)
        sys.exit(1)

    # --------------------------------------------------------
    # 4) Cargar JSON y validar estructura (debe ser lista)
    # --------------------------------------------------------
    with open(JSON_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)

    if not isinstance(data, list):
        print("ERROR: El JSON no es una lista de objetos.", file=sys.stderr)
        sys.exit(1)

    # --------------------------------------------------------
    # 5) Normalizar filas:
    #    - Asegurar las 4 columnas, con valores por defecto "N/A"
    #    - Mantener orden de columnas requerido para el Excel
    # --------------------------------------------------------
    rows = []
    for item in data:
        rows.append({
            "inventory_name": item.get("inventory_name", "N/A"),
            "hostname": item.get("hostname", "N/A"),
            "ip": item.get("ip", "N/A"),
            "os_full_version": item.get("os_full_version", "N/A"),
        })

    # --------------------------------------------------------
    # 6) Construir DataFrame en el orden deseado
    #    (opcional: ordenar por inventory_name y luego hostname)
    # --------------------------------------------------------
    df = pd.DataFrame(
        rows,
        columns=["inventory_name", "hostname", "ip", "os_full_version"]
    ).sort_values(by=["inventory_name", "hostname"], kind="stable", ignore_index=True)

    # --------------------------------------------------------
    # 7) Exportar a Excel con mejoras visuales:
    #    - Hoja: "Inventario SO"
    #    - Fila 1 congelada
    #    - Filtro auto en encabezados
    #    - Autoajuste ancho columnas
    # --------------------------------------------------------
    # Asegurar que OUT_DIR exista (si se ejecuta el script por separado)
    os.makedirs(OUT_DIR, exist_ok=True)

    with pd.ExcelWriter(XLSX_PATH, engine="openpyxl") as writer:
        sheet_name = "Inventario SO"
        df.to_excel(writer, index=False, sheet_name=sheet_name)

        # Tomar worksheet para formatear
        ws = writer.sheets[sheet_name]

        # Congelar la primera fila (encabezados)
        ws.freeze_panes = "A2"

        # Añadir autofiltro a toda la fila de encabezados
        last_col_letter = ws.cell(row=1, column=df.shape[1]).column_letter
        ws.auto_filter.ref = f"A1:{last_col_letter}{df.shape[0] + 1}"

        # Autoajustar columnas
        try:
            autosize_columns(ws, df)
        except Exception:
            # Si falla el autoajuste, no es crítico para la generación del archivo
            pass

    print(f"OK: Excel generado -> {XLSX_PATH}")

if __name__ == "__main__":
    main()
