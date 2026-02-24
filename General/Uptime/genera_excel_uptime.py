#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
genera_excel_uptime.py
- Lee un JSON con uptime por host (generado por Ansible)
- Genera un .xlsx con la información organizada
"""

import argparse
import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def autosize_columns(ws) -> None:
    """Ajusta ancho de columnas (aprox) según el contenido."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def safe_int(value, default=0) -> int:
    try:
        return int(value)
    except Exception:
        return default


def main() -> int:
    parser = argparse.ArgumentParser(description="Genera Excel de uptime desde JSON")
    parser.add_argument("--input", required=True, help="Ruta del JSON de entrada")
    parser.add_argument("--output", required=True, help="Ruta del XLSX de salida")
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)

    if not input_path.exists():
        raise FileNotFoundError(f"No existe el archivo de entrada: {input_path}")

    with input_path.open("r", encoding="utf-8") as f:
        data = json.load(f)

    if not isinstance(data, list):
        raise ValueError("El JSON debe ser una lista de objetos (list[dict]).")

    # Normaliza y ordena: uptime_days desc, luego inventory_hostname asc
    normalized = []
    for r in data:
        inv = (r.get("inventory_hostname") or "").strip()
        ahost = (r.get("ansible_hostname") or "").strip()
        up_sec = safe_int(r.get("uptime_seconds"), 0)
        up_days = safe_int(r.get("uptime_days"), up_sec // 86400)

        normalized.append({
            "inventory_hostname": inv,
            "ansible_hostname": ahost,
            "uptime_days": up_days,
            "uptime_seconds": up_sec,
        })

    normalized.sort(key=lambda x: (-x["uptime_days"], x["inventory_hostname"]))

    # Crea Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Uptime_BIA"

    headers = ["inventory_hostname", "ansible_hostname", "uptime_days", "uptime_seconds"]
    ws.append(headers)

    # Estilo encabezados
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Datos
    for r in normalized:
        ws.append([
            r["inventory_hostname"],
            r["ansible_hostname"],
            r["uptime_days"],
            r["uptime_seconds"],
        ])

    # Congelar primera fila y habilitar filtros
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    autosize_columns(ws)

    # Hoja de resumen
    ws2 = wb.create_sheet("Resumen")
    ws2.append(["Generado", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws2.append(["Total hosts", len(normalized)])

    if normalized:
        ws2.append(["Max uptime_days", max(r["uptime_days"] for r in normalized)])
        ws2.append(["Min uptime_days", min(r["uptime_days"] for r in normalized)])

    # Negrita en la primera columna del resumen
    for cell in ws2["A"]:
        cell.font = Font(bold=True)

    autosize_columns(ws2)

    # Crea carpeta destino si no existe (por si acaso)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb.save(output_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())