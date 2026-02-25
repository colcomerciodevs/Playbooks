#!/usr/bin/env python3
"""
curl_audit_to_excel.py

Lee un JSON consolidado (lista de dicts) generado por Ansible y produce un Excel
con formato para auditoría de curl/libcurl.

Incluye normalización robusta del campo affected_bundle para evitar casos donde
llegue como string ("True"/"False") y el Excel quede mal marcado.

Uso:
  python3 curl_audit_to_excel.py curl_audit.json curl_audit_report.xlsx
"""

import json
import sys
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# Orden de columnas en el Excel
COLUMNS = [
    "ip_from_inventory",
    "hostname",
    "os",

    "curl_installed",
    "curl_pkg_name",
    "libcurl_pkg_name",

    "curl_version",
    "curl_version_norm",
    "libcurl_version",

    "affected_range",
    "affected_bundle",

    "backend_ssl",
    "uses_gnutls",
    "uses_libssh",

    "supports_ldap",
    "supports_ldaps",
    "supports_sftp",
    "supports_scp",
    "supports_http3",

    "libcurl_whatrequires",

    "evidence_curl_version_line",
    "evidence_protocols_line",
    "evidence_features_line",
]

# Encabezados “bonitos”
HEADERS = {
    "ip_from_inventory": "IP (Inventario)",
    "hostname": "Hostname",
    "os": "OS",

    "curl_installed": "curl instalado",
    "curl_pkg_name": "Paquete curl detectado",
    "libcurl_pkg_name": "Paquete libcurl detectado",

    "curl_version": "curl (raw)",
    "curl_version_norm": "curl (X.Y.Z)",
    "libcurl_version": "libcurl (raw)",

    "affected_range": "Rango afectado",
    "affected_bundle": "AFECTADO (SI/NO)",

    "backend_ssl": "Backend TLS",
    "uses_gnutls": "Usa GnuTLS",
    "uses_libssh": "Usa libssh/libssh2",

    "supports_ldap": "Soporta LDAP",
    "supports_ldaps": "Soporta LDAPS",
    "supports_sftp": "Soporta SFTP",
    "supports_scp": "Soporta SCP",
    "supports_http3": "Soporta HTTP/3",

    "libcurl_whatrequires": "Paquetes que requieren libcurl (whatrequires)",

    "evidence_curl_version_line": "Evidencia: curl --version (línea 1)",
    "evidence_protocols_line": "Evidencia: Protocols",
    "evidence_features_line": "Evidencia: Features",
}


def read_json(path: Path) -> List[Dict[str, Any]]:
    """Lee el JSON y valida que sea una lista de dicts."""
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, list):
        raise ValueError("El JSON esperado debe ser una LISTA de registros por host.")
    # Validación suave: cada item debe ser dict
    for i, item in enumerate(data):
        if not isinstance(item, dict):
            raise ValueError(f"El item #{i} del JSON no es un dict.")
    return data


def normalize_bool(value: Any) -> Any:
    """
    Normaliza valores booleanos que pueden venir como:
      - bool True/False
      - string "true"/"false"/"si"/"no"/"yes"/"0"/"1"
      - int 0/1

    Retorna:
      - True/False si logra interpretarlo
      - el valor original si no puede normalizarlo
    """
    if isinstance(value, bool):
        return value

    if isinstance(value, int):
        if value == 1:
            return True
        if value == 0:
            return False
        return value

    if isinstance(value, str):
        s = value.strip().lower()
        if s in ("true", "yes", "y", "si", "sí", "1"):
            return True
        if s in ("false", "no", "n", "0"):
            return False
        return value

    return value


def autosize_columns(ws) -> None:
    """Ajusta ancho de columnas aproximado con límites razonables."""
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        max_len = len(HEADERS.get(col_name, col_name))
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            v = row[0].value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)

        # límites para columnas largas
        if col_name in (
            "libcurl_whatrequires",
            "evidence_protocols_line",
            "evidence_features_line",
            "evidence_curl_version_line",
        ):
            width = min(max_len, 80)
        else:
            width = min(max_len, 40)

        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, width + 2)


def main() -> int:
    if len(sys.argv) != 3:
        print("Uso: python3 curl_audit_to_excel.py <input.json> <output.xlsx>", file=sys.stderr)
        return 2

    input_json = Path(sys.argv[1])
    output_xlsx = Path(sys.argv[2])

    if not input_json.exists():
        print(f"ERROR: No existe el archivo JSON: {input_json}", file=sys.stderr)
        return 2

    records = read_json(input_json)

    wb = Workbook()
    ws = wb.active
    ws.title = "curl_audit"

    # Estilos
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")  # azul claro
    wrap = Alignment(wrap_text=True, vertical="top")
    top_left = Alignment(vertical="top", horizontal="left", wrap_text=True)

    # Encabezados
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=HEADERS.get(col_name, col_name))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

    # Congelar encabezado y autofiltro
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # Colores para afectación
    fill_affected = PatternFill("solid", fgColor="FFC7CE")      # rojo claro
    fill_not_affected = PatternFill("solid", fgColor="C6EFCE")  # verde claro
    fill_unknown = PatternFill("solid", fgColor="FFEB9C")       # amarillo claro

    affected_col = COLUMNS.index("affected_bundle") + 1

    # Escribir filas
    for row_idx, rec in enumerate(records, start=2):
        # Volcar celdas (raw)
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            val = rec.get(col_name, "")
            ws.cell(row=row_idx, column=col_idx, value=val).alignment = top_left

        # Normalizar y pintar la celda AFECTADO (SI/NO)
        raw_affected = rec.get("affected_bundle", "")
        affected_value = normalize_bool(raw_affected)
        affected_cell = ws.cell(row=row_idx, column=affected_col)

        if isinstance(affected_value, bool):
            affected_cell.value = "SI" if affected_value else "NO"
            affected_cell.font = Font(bold=True)
            affected_cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            affected_cell.fill = fill_affected if affected_value else fill_not_affected
        else:
            # Si viene vacío o no interpretable, marcamos N/A y amarillo
            affected_cell.value = "N/A" if (affected_value == "" or affected_value is None) else str(affected_value)
            affected_cell.font = Font(bold=True)
            affected_cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            affected_cell.fill = fill_unknown

        # Mejor lectura en campos largos
        long_fields = [
            "libcurl_whatrequires",
            "evidence_curl_version_line",
            "evidence_protocols_line",
            "evidence_features_line",
        ]
        for lf in long_fields:
            lf_col = COLUMNS.index(lf) + 1
            ws.cell(row=row_idx, column=lf_col).alignment = wrap

    autosize_columns(ws)

    # Guardar
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)

    print(f"OK: Excel generado en: {output_xlsx}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())