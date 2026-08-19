#!/usr/bin/env python3
# ============================================================
# Script: convert_json_to_excel.py
# Objetivo:
#   Leer Salidas_Playbooks/inventario_servers.json (lista de objetos)
#   y generar un Excel con columnas:
#     - inventory_name, hostname, ip, os_full_version
#     - serial, modelo            (dmidecode)
#     - procesador, cpu_total     (lscpu)
#     - memoria, memoria_gb       (free -h)
#     - disco_total, disco_gb, disco_detalle  (pvs)
#
# Limpieza extra:
#   - Elimina saltos de línea y dobles espacios
#   - Quita espacios al inicio/fin en todas las celdas de texto
#   - Convierte memoria y disco a valor numérico (GB) para poder sumar/ordenar
#   - Ordena las IPs de forma natural (no alfabética)
#
# Requisitos: pandas, openpyxl
# Uso:
#   python3 scripts/convert_json_to_excel.py
# ============================================================

import ipaddress
import json
import os
import re
import sys
from datetime import datetime

try:
    import pandas as pd
except Exception:
    print("ERROR: Se requiere 'pandas' instalado. Instale con: pip install pandas openpyxl", file=sys.stderr)
    sys.exit(2)

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.join(BASE_DIR, "Salidas_Playbooks")
JSON_PATH = os.path.join(OUT_DIR, "inventario_servers.json")
TS = datetime.now().strftime("%Y%m%d_%H%M%S")
XLSX_PATH = os.path.join(OUT_DIR, f"inventario_servers_{TS}.xlsx")

# Orden y encabezados "bonitos" para el Excel
COLUMNS = [
    ("inventory_name",  "Nombre Inventario"),
    ("hostname",        "Hostname S.O."),
    ("ip",              "IP"),
    ("os_full_version", "Versión S.O."),
    ("serial",          "Serial"),
    ("modelo",          "Fabricante / Modelo"),
    ("procesador",      "Procesador"),
    ("cpu_total",       "CPUs"),
    ("memoria",         "Memoria"),
    ("memoria_gb",      "Memoria (GB)"),
    ("disco_total",     "Disco LVM"),
    ("disco_gb",        "Disco LVM (GB)"),
    ("disco_detalle",   "Detalle PVs"),
]

# Valores que dmidecode devuelve cuando no hay serial real
SERIAL_INVALIDOS = {
    "", "na", "n/a", "none", "null", "not specified", "not available",
    "system serial number", "to be filled by o.e.m.", "0", "unknown",
}


# --- Limpiador de texto seguro y uniforme ---
def clean_text(value, default="N/A"):
    if value is None:
        return default
    s = str(value)
    # Normaliza saltos de línea/tabs y colapsa espacios
    s = s.replace("\r", " ").replace("\n", " ").replace("\t", " ")
    s = " ".join(s.split())
    s = s.strip().strip('"').strip()
    return s if s else default


def clean_serial(value, default="N/A"):
    s = clean_text(value, default="")
    return default if s.lower() in SERIAL_INVALIDOS else s


def to_gb(value):
    """Convierte '31Gi', '7.6G', '512Mi', '1.5Ti', '250.00 GB' -> float en GB."""
    s = clean_text(value, default="")
    if not s:
        return None
    m = re.search(r"([\d]+[.,]?[\d]*)\s*([KMGTP])?i?B?", s, re.IGNORECASE)
    if not m:
        return None
    try:
        num = float(m.group(1).replace(",", "."))
    except ValueError:
        return None
    unidad = (m.group(2) or "G").upper()
    factor = {"K": 1 / 1048576, "M": 1 / 1024, "G": 1, "T": 1024, "P": 1048576}
    return round(num * factor.get(unidad, 1), 2)


def to_int(value):
    s = clean_text(value, default="")
    m = re.search(r"\d+", s)
    return int(m.group(0)) if m else None


def ip_sort_key(value):
    """Ordena IPs numéricamente; lo que no sea IP va al final."""
    try:
        return (0, int(ipaddress.ip_address(str(value).strip())))
    except Exception:
        return (1, 0)


def autosize_columns(ws, df, padding=2, max_width=60):
    from openpyxl.utils import get_column_letter
    for idx, col in enumerate(df.columns, start=1):
        header_len = len(str(col))
        body_len = max((len(str(x)) for x in df[col].astype(str)), default=0)
        width = min(max(header_len, body_len) + padding, max_width)
        ws.column_dimensions[get_column_letter(idx)].width = width


def format_header(ws, ncols):
    from openpyxl.styles import Alignment, Font, PatternFill
    fill = PatternFill("solid", fgColor="1F4E78")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22


def main():
    if not os.path.exists(JSON_PATH):
        print(f"ERROR: No existe el archivo JSON esperado: {JSON_PATH}", file=sys.stderr)
        sys.exit(1)

    with open(JSON_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)

    if not isinstance(data, list):
        print("ERROR: El JSON no es una lista de objetos.", file=sys.stderr)
        sys.exit(1)

    # Normaliza cada fila (limpieza aplicada a todas las columnas)
    rows = []
    for item in data:
        if not isinstance(item, dict):
            continue
        memoria = clean_text(item.get("memoria"))
        disco = clean_text(item.get("disco_total"))
        rows.append({
            "inventory_name":  clean_text(item.get("inventory_name")),
            "hostname":        clean_text(item.get("hostname")),
            "ip":              clean_text(item.get("ip")),
            "os_full_version": clean_text(item.get("os_full_version")),
            "serial":          clean_serial(item.get("serial")),
            "modelo":          clean_text(item.get("modelo")),
            "procesador":      clean_text(item.get("procesador")),
            "cpu_total":       to_int(item.get("cpu_total")),
            "memoria":         memoria,
            "memoria_gb":      to_gb(memoria),
            "disco_total":     disco,
            "disco_gb":        to_gb(disco),
            "disco_detalle":   clean_text(item.get("disco_detalle")),
        })

    if not rows:
        print("ADVERTENCIA: El JSON no contiene registros válidos.", file=sys.stderr)

    campos = [c[0] for c in COLUMNS]
    df = pd.DataFrame(rows, columns=campos)

    # Orden natural por IP y luego por nombre de inventario
    df["_ipkey"] = df["ip"].map(ip_sort_key)
    df = df.sort_values(by=["_ipkey", "inventory_name"], kind="stable", ignore_index=True)
    df = df.drop(columns=["_ipkey"])

    # Hoja resumen: cantidad de equipos por versión de S.O.
    resumen = (
        df.groupby("os_full_version", dropna=False)
          .size()
          .reset_index(name="cantidad")
          .sort_values(by=["cantidad", "os_full_version"], ascending=[False, True], ignore_index=True)
    )
    resumen.columns = ["Versión S.O.", "Cantidad de equipos"]

    # Encabezados legibles
    df_out = df.rename(columns=dict(COLUMNS))

    os.makedirs(OUT_DIR, exist_ok=True)

    with pd.ExcelWriter(XLSX_PATH, engine="openpyxl") as writer:
        sheet_name = "Inventario SO"
        df_out.to_excel(writer, index=False, sheet_name=sheet_name)

        ws = writer.sheets[sheet_name]
        ws.freeze_panes = "A2"

        last_col_letter = ws.cell(row=1, column=df_out.shape[1]).column_letter
        ws.auto_filter.ref = f"A1:{last_col_letter}{df_out.shape[0] + 1}"

        try:
            from openpyxl.styles import Font
            for fila in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for celda in fila:
                    celda.font = Font(name="Arial")
            format_header(ws, df_out.shape[1])
            autosize_columns(ws, df_out)
        except Exception:
            pass

        # Segunda hoja con el resumen
        resumen.to_excel(writer, index=False, sheet_name="Resumen SO")
        ws2 = writer.sheets["Resumen SO"]
        ws2.freeze_panes = "A2"
        try:
            from openpyxl.styles import Font
            for fila in ws2.iter_rows(min_row=2, max_row=ws2.max_row, max_col=ws2.max_column):
                for celda in fila:
                    celda.font = Font(name="Arial")
            format_header(ws2, resumen.shape[1])
            autosize_columns(ws2, resumen)
        except Exception:
            pass

    print(f"OK: {len(df_out)} equipos procesados")
    print(f"OK: Excel generado -> {XLSX_PATH}")


if __name__ == "__main__":
    main()