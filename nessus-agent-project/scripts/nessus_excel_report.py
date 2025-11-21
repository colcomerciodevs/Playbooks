#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generador de reporte Excel desde el JSON de resultados de despliegue Nessus Agent.

- Lee un JSON (lista de dicts) producido por Ansible con la clave `nessus_result` por host.
- Crea un Excel con columnas útiles para auditoría (incluye "Link Status").
- Aplica estilos básicos, autoajuste de columnas y filtros.

Uso:
    python3 nessus_excel_report.py <input_json> <output_xlsx>

Ejemplo:
    python3 nessus_excel_report.py ./Salidas_Playbooks/nessus_results.json ./Salidas_Playbooks/nessus_agent_reporte.xlsx
"""

import sys
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# 1) Validar argumentos
if len(sys.argv) != 3:
    print("Uso: python3 nessus_excel_report.py <input_json> <output_xlsx>")
    sys.exit(1)

json_file = sys.argv[1]
xlsx_file = sys.argv[2]
os.makedirs(os.path.dirname(os.path.abspath(xlsx_file)), exist_ok=True)

# 2) Leer JSON
try:
    with open(json_file, "r", encoding="utf-8") as f:
        data = json.load(f)
except FileNotFoundError:
    print(f"❌ No se encontró el archivo: {json_file}")
    sys.exit(2)
except json.JSONDecodeError as e:
    print(f"❌ Error al decodificar JSON: {e}")
    sys.exit(3)

if not isinstance(data, list):
    print("❌ El JSON no contiene una lista de resultados por host.")
    sys.exit(4)

# 3) Ordenar por host (opcional)
try:
    data_sorted = sorted(data, key=lambda h: (h.get("host") or ""))
except Exception:
    data_sorted = data

# 4) Crear libro y hoja
wb = Workbook()
ws = wb.active
ws.title = "Reporte Nessus Agent"

# 5) Encabezados y estilos
headers = [
    "Host",
    "Hostname",  
    "IP (Inventario)",    
    "OS Name",
    "OS Family",
    "Versión",
    "Agent Version",
    "Agente Instalado",
    "Servicio Activo",
    "Vinculado",
    "Link Status",
    "Manager Host",
    "Manager Port",
    "Grupos",
    "Link Output",
    "Error",
    "Timestamp",
]
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
center_align = Alignment(horizontal="center", vertical="center")

ws.append(headers)
for col_idx in range(1, len(headers) + 1):
    c = ws.cell(row=1, column=col_idx)
    c.fill = header_fill
    c.font = header_font
    c.alignment = center_align

# 6) Helper para valores vacíos
def nz(val, default="N/A"):
    v = (val or "")
    v = v.strip() if isinstance(v, str) else v
    return v if v not in ("", None) else default

# 7) Volcar filas
for host in data_sorted:
    # Fallback para link status si no vino calculado
    computed_link_status = host.get("link_status")
    if not computed_link_status:
        computed_link_status = "Connected to Manager" if host.get("linked") else "Not linked to a manager"

    row = [
        nz(host.get("host"), ""),
        nz(host.get("hostname") or host.get("ansible_hostname"), ""),
        nz(host.get("ip_address")),             
        nz(host.get("os_name")),
        nz(host.get("os_family")),
        nz(host.get("os_version")),
        nz(host.get("agent_version")),
        "✅" if host.get("installed") else "❌",
        "✅" if host.get("service_active") else "❌",
        "✅" if host.get("linked") else "❌",
        nz(computed_link_status),
        nz(host.get("manager_host")),
        nz(host.get("manager_port")),
        nz(host.get("groups")),
        nz(host.get("link_output")),
        nz(host.get("error")),
        nz(host.get("timestamp")),
    ]
    ws.append(row)

# 8) Estética: ajustar columnas y envolver texto en columnas largas
wrap_align = Alignment(wrap_text=True, vertical="top")
long_cols = {"Link Output", "Error"}
for col_idx, title in enumerate(headers, start=1):
    if title in long_cols:
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=col_idx).alignment = wrap_align

for col in ws.columns:
    max_len = 0
    letter = col[0].column_letter
    for cell in col:
        if cell.value:
            max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions[letter].width = min(max_len + 2, 60)

# 9) Usabilidad
ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

# 10) Guardar
try:
    wb.save(xlsx_file)
    print(f"✅ Reporte Excel generado correctamente: {xlsx_file}")
except Exception as e:
    print(f"❌ Error al guardar Excel: {e}")
    sys.exit(5)
