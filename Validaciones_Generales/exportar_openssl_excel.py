#!/usr/bin/env python3
import sys
import os
import json
import openpyxl
import re
from packaging import version
from openpyxl.styles import PatternFill

# Verificar argumento de ruta
if len(sys.argv) != 2:
    print("Uso: python3 exportar_openssl_excel.py <directorio_salida>")
    sys.exit(1)

base_dir = sys.argv[1]
json_path = os.path.join(base_dir, "openssl_consolidado.json")

# CVEs, sus rangos de versiones afectadas y severidades
CVE_RANGES = {
    "CVE-2022-1292": [("1.0.2", "1.0.2ze"), ("1.1.1", "1.1.1o"), ("3.0.0", "3.0.3")],
    "CVE-2022-2068": [("1.0.2", "1.0.2zf"), ("1.1.1", "1.1.1p"), ("3.0.0", "3.0.4")],
    "CVE-2022-4304": [("1.0.2", "1.0.2zg"), ("1.1.1", "1.1.1t"), ("3.0.0", "3.0.8")],
    "CVE-2023-0215": [("1.0.2", "1.0.2zg"), ("1.1.1", "1.1.1t"), ("3.0.0", "3.0.8")],
    "CVE-2023-0286": [("1.0.2", "1.0.2zg"), ("1.1.1", "1.1.1t"), ("3.0.0", "3.0.8")],
    "CVE-2023-0464": [("1.0.2", "1.0.2zh"), ("1.1.1", "1.1.1u"), ("3.0.0", "3.0.9"), ("3.1.0", "3.1.1")],
    "CVE-2023-0465": [("1.0.2", "1.0.2zh"), ("1.1.1", "1.1.1u"), ("3.0.0", "3.0.9"), ("3.1.0", "3.1.1")],
    "CVE-2023-0466": [("1.0.2", "1.0.2zh"), ("1.1.1", "1.1.1u"), ("3.0.0", "3.0.9"), ("3.1.0", "3.1.1")],
    "CVE-2023-3446": [("1.0.2", "1.0.2zi"), ("1.1.1", "1.1.1v"), ("3.0.0", "3.0.10"), ("3.1.0", "3.1.2")],
    "CVE-2023-3817": [("1.0.2", "1.0.2zi"), ("1.1.1", "1.1.1v"), ("3.0.0", "3.0.10"), ("3.1.0", "3.1.2")],
}

CVE_SEVERITY = {
    "CVE-2022-1292": "Medium",
    "CVE-2022-2068": "High",
    "CVE-2022-4304": "High",
    "CVE-2023-0215": "High",
    "CVE-2023-0286": "Medium",
    "CVE-2023-0464": "High",
    "CVE-2023-0465": "High",
    "CVE-2023-0466": "High",
    "CVE-2023-3446": "High",
    "CVE-2023-3817": "High",
}

MIN_SEGURO = {
    "1.0.2": "1.0.2zi",
    "1.1.1": "1.1.1v",
    "3.0.0": "3.0.10",
    "3.0": "3.0.10",
    "3.1.0": "3.1.2",
    "3.1": "3.1.2"
}

# Convierte una versión OpenSSL con sufijo a versión comparable numérica
def openssl_version_to_number(ver_str):
    match = re.search(r'OpenSSL\s+(\d+\.\d+\.\d+)([a-z]*)', ver_str, re.IGNORECASE)
    if not match:
        return None
    base = match.group(1)
    suffix = match.group(2).lower()
    if suffix:
        suffix_value = 0
        for i, ch in enumerate(suffix[::-1]):
            suffix_value += (ord(ch) - ord('a') + 1) * (26 ** i)
        return f"{base}.{suffix_value}"
    else:
        return base

# Evalúa si la versión actual está dentro del rango afectado
def version_in_range(current, start, end):
    try:
        current_v = version.parse(current)
        start_v = version.parse(openssl_version_to_number(f"OpenSSL {start}"))
        end_v = version.parse(openssl_version_to_number(f"OpenSSL {end}"))
        return start_v <= current_v < end_v
    except Exception:
        return False

# Evalúa si requiere actualización (devuelve versión segura si aplica)
def requiere_actualizacion(parsed_version):
    if not parsed_version:
        return "Desconocido"
    try:
        parsed = version.parse(parsed_version)
        base_parts = parsed.base_version.split(".")
        for i in range(3, 0, -1):
            ref_base = ".".join(base_parts[:i])
            if ref_base in MIN_SEGURO:
                min_segura = MIN_SEGURO[ref_base]
                min_segura_parsed = version.parse(openssl_version_to_number(f"OpenSSL {min_segura}"))
                if parsed < min_segura_parsed:
                    return min_segura
                else:
                    return "No"
        return "Desconocido"
    except:
        return "Desconocido"

# Evalúa vulnerabilidad por versión
def evaluar_vulnerabilidad(ver_str):
    parsed_version = openssl_version_to_number(ver_str)
    if not parsed_version:
        return "No evaluado", [], "Formato no reconocido", parsed_version

    cves_afectan = []
    try:
        for cve, ranges in CVE_RANGES.items():
            for start, end in ranges:
                if version_in_range(parsed_version, start, end):
                    cves_afectan.append(cve)
                    break
        if cves_afectan:
            return "Vulnerable", cves_afectan, "", parsed_version
        else:
            return "No vulnerable", [], "", parsed_version
    except Exception as e:
        return "No evaluado", [], str(e), parsed_version

# Leer JSON de entrada
with open(json_path, "r") as f:
    data = json.load(f)

# Crear archivo Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.append(["IP", "Versión OpenSSL", "Versión Normalizada", "Estado", "CVEs relacionados", "Severidad", "¿Requiere actualización?", "Motivo (si no evaluado)"])

# Colores para columna estado
fill_vulnerable = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
fill_no_vuln = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

# Contadores
vulnerables = 0
no_vulnerables = 0
no_evaluados = 0

# Procesar cada host
for item in data:
    ip = item.get("ip", "desconocido")
    version_str = item.get("version", "desconocido")
    estado, cves, motivo, parsed_version = evaluar_vulnerabilidad(version_str)
    requiere_update = requiere_actualizacion(parsed_version)

    severidades = sorted(set([CVE_SEVERITY.get(cve, "Desconocida") for cve in cves]))

    if estado == "Vulnerable":
        vulnerables += 1
    elif estado == "No vulnerable":
        no_vulnerables += 1
    else:
        no_evaluados += 1

    row = [
        ip,
        version_str,
        parsed_version,
        estado,
        ", ".join(cves),
        ", ".join(severidades),
        requiere_update,
        motivo
    ]
    ws.append(row)

    # Colorear solo la columna del estado
    estado_cell = ws.cell(row=ws.max_row, column=4)
    if estado == "Vulnerable":
        estado_cell.fill = fill_vulnerable
    elif estado == "No vulnerable":
        estado_cell.fill = fill_no_vuln

# Guardar Excel
excel_path = os.path.join(base_dir, "reporte_openssl_vulnerabilidades.xlsx")
wb.save(excel_path)

# Mostrar resumen en consola
total = len(data)
print(f"\n[✓] Excel generado: {excel_path}")
print(f"Total evaluados: {total}")
print(f" - Vulnerables     : {vulnerables}")
print(f" - No vulnerables  : {no_vulnerables}")
print(f" - No evaluados    : {no_evaluados}")
