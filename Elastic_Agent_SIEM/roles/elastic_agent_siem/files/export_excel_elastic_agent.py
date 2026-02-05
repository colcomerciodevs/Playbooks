#!/usr/bin/env python3
"""
Exporta un reporte Excel (XLSX) a partir de un JSON consolidado con datos por host.

Soporta 3 escenarios:
1) SOLO RE-ENROLL   (record_min.yml): incluye reenroll_* y NO valida servicio
2) SOLO RESTART     (service.yml): incluye restart/status/ok/version
3) RE-ENROLL+RESTART: incluye ambos (reenroll_* + restart/status/ok/version)

Columnas (orden):
1) Inventario (inventory_hostname)          -> inventory_name
2) Hostname real (ansible_hostname)         -> hostname
3) IP del inventario (ansible_host)         -> inventory_ip
4) Versión Elastic Agent                    -> elastic_version
5) Re-enroll ejecutado                      -> reenroll_ran
6) Re-enroll OK                             -> reenroll_ok
7) Mensaje Re-enroll                        -> reenroll_msg
8) Reinicio OK                              -> restart_ok
9) Estado del servicio (is-active)          -> is_active
10) OK (verde / rojo)                       -> ok
11) Extracto de status                      -> status_excerpt

Uso:
  export_excel_elastic_agent.py <input.json> <output.xlsx>

Requisitos (en el nodo controlador):
  python3 -m pip install --user openpyxl
"""

import json
import sys
from typing import Any, Dict, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Colores (suaves) para semáforo
GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
HEADER = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


def autosize(ws, max_width: int = 90) -> None:
    """Auto-ajuste de ancho de columnas con límite."""
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        max_len = 0
        for cell in ws[col_letter]:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def safe_bool(val: Any) -> bool:
    """Convierte valores típicos a booleano."""
    if isinstance(val, bool):
        return val
    if val is None:
        return False
    if isinstance(val, str):
        return val.strip().lower() in ("1", "true", "yes", "si", "sí", "ok")
    if isinstance(val, (int, float)):
        return val != 0
    return False


def norm_str(val: Any, default: str = "N/A") -> str:
    """Normaliza strings: None/vacío -> default."""
    if val is None:
        return default
    s = str(val).strip()
    return s if s else default


def load_json(path: str) -> List[Dict[str, Any]]:
    """Carga JSON y valida que sea lista de dicts."""
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except FileNotFoundError:
        raise SystemExit(f"No existe el JSON de entrada: {path}")
    except json.JSONDecodeError as e:
        raise SystemExit(f"JSON inválido: {path} -> {e}")

    if not isinstance(data, list):
        raise SystemExit("El JSON debe ser una lista de registros por host.")

    # Filtrar solo dicts
    out: List[Dict[str, Any]] = []
    for item in data:
        if isinstance(item, dict):
            out.append(item)
    return out


def main() -> int:
    if len(sys.argv) != 3:
        print("Uso: export_excel_elastic_agent.py <input.json> <output.xlsx>", file=sys.stderr)
        return 2

    in_json = sys.argv[1]
    out_xlsx = sys.argv[2]

    data = load_json(in_json)

    wb = Workbook()
    ws = wb.active
    ws.title = "Elastic Agent"

    headers = [
        "Inventario (inventory_hostname)",
        "Hostname real (ansible_hostname)",
        "IP del inventario",
        "Versión Elastic Agent",

        "Re-enroll ejecutado",
        "Re-enroll OK",
        "Mensaje re-enroll",

        "Reinicio OK",
        "Estado del servicio (is-active)",
        "OK (verde/rojo)",
        "Extracto de status",
    ]
    ws.append(headers)

    # Estilo encabezado
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Filas
    for item in data:
        inventory_name = norm_str(item.get("inventory_name"), "N/A")
        hostname = norm_str(item.get("hostname"), "N/A")
        inventory_ip = norm_str(item.get("inventory_ip"), "N/A")

        # Nunca permitir None/vacío en versión
        elastic_version = norm_str(item.get("elastic_version"), "N/A")

        # Re-enroll (puede venir de record_min.yml o service.yml si preservas campos)
        reenroll_ran = safe_bool(item.get("reenroll_ran", False))
        reenroll_ok = safe_bool(item.get("reenroll_ok", False))
        reenroll_msg = norm_str(item.get("reenroll_msg"), "N/A")

        # Restart/Status (puede venir vacío en solo reenroll)
        restart_ok = safe_bool(item.get("restart_ok", False))
        is_active = norm_str(item.get("is_active"), "unknown")
        ok = safe_bool(item.get("ok", False))
        excerpt = norm_str(item.get("status_excerpt"), "")

        ws.append([
            inventory_name,
            hostname,
            inventory_ip,
            elastic_version,

            "SI" if reenroll_ran else "NO",
            "OK" if reenroll_ok else "NO OK",
            reenroll_msg,

            "SI" if restart_ok else "NO",
            is_active,
            "OK" if ok else "NO OK",
            excerpt,
        ])

        row = ws.max_row

        # -----------------------------
        # Colores / Semáforos
        # -----------------------------

        # OK global (col 10) y Estado (col 9) dependen de "ok"
        fill_ok = GREEN if ok else RED
        ws.cell(row=row, column=10).fill = fill_ok               # OK global
        ws.cell(row=row, column=9).fill = fill_ok                # Estado servicio

        # Reinicio OK (col 8)
        ws.cell(row=row, column=8).fill = GREEN if restart_ok else RED

        # Re-enroll OK (col 6) solo si se ejecutó re-enroll
        if reenroll_ran:
            ws.cell(row=row, column=6).fill = GREEN if reenroll_ok else RED
        else:
            # Si no se ejecutó, lo dejamos neutro (amarillo suave)
            ws.cell(row=row, column=6).fill = YELLOW

        # Versión: si NO INSTALADO -> rojo (útil en auditoría)
        if elastic_version.upper() == "NO INSTALADO":
            ws.cell(row=row, column=4).fill = RED

        # -----------------------------
        # Alineación
        # -----------------------------
        for col in [1, 2, 3, 4, 5, 6, 8, 9, 10]:
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")

        # Mensaje re-enroll y extracto con wrap
        ws.cell(row=row, column=7).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=11).alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    autosize(ws, max_width=90)
    wb.save(out_xlsx)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
